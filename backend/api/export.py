import io
from io import BytesIO
from datetime import date, datetime, timedelta
from typing import Optional
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
)
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.orm import joinedload
from sqlalchemy import func
from database.connection import get_db
from database.models import (
    Transaction,
    GlAccount,
    Settlement,
    AdvanceRequest,
    PrintedCheck,
    Vendor,
)
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak
)

from api.advance import update_ppc_status
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import A4

from api.dashboard import (
    DashboardSource,
    get_transaction_model
)

router = APIRouter(
    prefix="/export",
    tags=["Export"]
)

# HELPER
def apply_date_filter(
        query,
        model,
        start_date,
        end_date
):
    if start_date:
        query = query.filter(
            model.posting_date >= start_date
        )
    if end_date:
        query = query.filter(
            model.posting_date <= end_date
        )
    return query

def rupiah(value):
    return f"Rp {float(value or 0):,.0f}"

def rupiah_parts(value):

    val = float(value or 0)
    angka = f"{val:,.0f}".replace(",", ".")
    return "Rp.", angka

def format_period(start_date, end_date):
    if not start_date and not end_date:
        return "All Data"
    if start_date and end_date:
        return f"{start_date} s/d {end_date}"
    if start_date:
        return f"{start_date} s/d Sekarang"
    return f"Sampai {end_date}"

def count_workdays(start, end):
    days = 0
    current = start
    while current <= end:
        if current.weekday() < 5:
            days += 1
        current += timedelta(days=1)
    return days

# PDF HELPER
def calculate_percentage(value, total):

    if not total:
        return 0.0

    return (float(value) / float(total)) * 100


def calculate_growth(current, previous):
    """
    Growth dibanding periode sebelumnya.
    """
    if previous in (0, None):
        return None

    return round(
        ((current - previous) / previous) * 100,
        2
    )

def determine_trend_group(start_date, end_date):
    """
    <=31 hari  -> day

    >31 hari   -> month

    tanpa filter -> month
    """

    if not start_date or not end_date:
        return "month"

    total_days = (end_date - start_date).days

    if total_days <= 31:
        return "day"

    return "month"


def create_table(data, widths=None, extra_style=None):
    table = Table(
        data,
        colWidths=widths
    )
    table.setStyle(TABLE_STYLE)
    if extra_style:
        table.setStyle(TableStyle(extra_style))
    return table

# Advance Helper
def apply_advance_date_filter(
    query,
    start_date,
    end_date
):

    if start_date:
        query = query.filter(
            AdvanceRequest.request_date >= start_date
        )

    if end_date:
        query = query.filter(
            AdvanceRequest.request_date <= end_date
        )

    return query

# EXCEL FORMAT HELPER
EXCEL_DATE_FORMAT = "dd/mm/yyyy"
EXCEL_CURRENCY_FORMAT = '#,##0'

def format_excel_worksheet(
    worksheet,
    date_columns=None,
    currency_columns=None
):
    date_columns = date_columns or []
    currency_columns = currency_columns or []

    # Header
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Mapping nama header -> nomor kolom
    header_map = {}

    for cell in worksheet[1]:
        if cell.value:
            header_map[str(cell.value)] = cell.column

    # Format cells (number formats and alignments)
    for header_name, column_index in header_map.items():
        is_currency = header_name in currency_columns
        is_date = header_name in date_columns

        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(
                row=row,
                column=column_index
            )

            if is_currency:
                if cell.value is not None:
                    cell.number_format = EXCEL_CURRENCY_FORMAT
                cell.alignment = Alignment(
                    horizontal="right"
                )
            else:
                if is_date and cell.value is not None:
                    cell.number_format = EXCEL_DATE_FORMAT
                cell.alignment = Alignment(
                    horizontal="left"
                )

    # Auto width
    for column_cells in worksheet.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                value = str(cell.value)
                max_length = max(
                    max_length,
                    len(value)
                )

        column_letter = get_column_letter(
            column_cells[0].column
        )

        worksheet.column_dimensions[
            column_letter
        ].width = max(
            max_length + 3,
            15
        )

TABLE_STYLE = TableStyle([

    # Header
    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ("FONTSIZE", (0, 0), (-1, 0), 9),
    ("ALIGN", (0, 0), (-1, 0), "CENTER"),

    # Isi tabel
    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
    ("FONTSIZE", (0, 1), (-1, -1), 8),

    # Kolom pertama selalu rata kiri
    ("ALIGN", (0, 0), (0, -1), "LEFT"),
    # Kolom kedua dst rata kiri kecuali kolom angka/persen
    ("ALIGN", (1, 1), (1, -1), "LEFT"),
    # Kolom dari kolom ke-3 rata kanan (angka/persen)
    ("ALIGN", (2, 1), (-1, -1), "RIGHT"),

    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

    # Grid
    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),

    # Padding
    ("TOPPADDING", (0, 0), (-1, -1), 5),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ("RIGHTPADDING", (0, 0), (-1, -1), 6),

    # Warna baris ganjil/genap
    ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
        colors.white,
        colors.HexColor("#F0F4FA")
    ]),

])

# EXPORT EXCEL
@router.get("/excel")
def export_excel(

    source: DashboardSource = DashboardSource.rungkut,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)

):

    try:

        TransactionModel = get_transaction_model(
            source
        )

        query = db.query(
            TransactionModel
        )

        query = apply_date_filter(
            query,
            TransactionModel,
            start_date,
            end_date
        )

        transactions = (
            query
            .order_by(
                TransactionModel.posting_date
            )
            .all()
        )

        rows = []

        for row in transactions:

            rows.append({

                "Posting Date":
                    row.posting_date,

                "Document No":
                    row.document_no,

                "Amount":
                    row.amount,

                "Currency":
                    row.currency,

                "GL Account":
                    row.gl_account,

                "Cost Center":
                    row.cost_center,

                "Reference":
                    row.reference,

                "Transaction Type":
                    row.transaction_type,

                "Description":
                    row.description,

                "Month":
                    row.month,

                "Year":
                    row.year,

                "Uploaded At":
                    row.uploaded_at.strftime(
                        "%Y-%m-%d %H:%M:%S"
                    )
                    if row.uploaded_at
                    else ""

            })

        df = pd.DataFrame(rows)

        # Jika tidak ada data
        if df.empty:

            df = pd.DataFrame(columns=[

                "Posting Date",
                "Document No",
                "Amount",
                "Currency",
                "GL Account",
                "Cost Center",
                "Reference",
                "Transaction Type",
                "Description",
                "Month",
                "Year",
                "Uploaded At"

            ])

        output = io.BytesIO()

        with pd.ExcelWriter(
            output,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="Transactions",
                index=False
            )

            worksheet = writer.sheets["Transactions"]

            format_excel_worksheet(
                worksheet,
                date_columns=[
                    "Posting Date"
                ],
                currency_columns=[
                    "Amount"
                ]
            )

        output.seek(0)

        filename = (
            f"Export Dashboard_{source.value.title()}.xlsx"
        )

        return StreamingResponse(

            output,

            media_type=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),

            headers={

                "Content-Disposition":
                f'attachment; filename="{filename}"'

            }

        )

    except Exception as e:

        raise HTTPException(

            status_code=500,
            detail=str(e)

        )

# EXPORT PDF
@router.get("/pdf")
def export_dashboard_pdf(
    source: DashboardSource = DashboardSource.rungkut,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    trend_group: str = Query("month"),
    db: Session = Depends(get_db)

):
    try:
        TransactionModel = get_transaction_model(
        source
    )
        # SUMMARY
        summary_query = apply_date_filter(
            db.query(
                TransactionModel
            ),
            TransactionModel,
            start_date,
            end_date
        )

        summary = summary_query.with_entities(
            func.sum(TransactionModel.amount),
            func.count(TransactionModel.id),
            func.count(
                func.distinct(
                    TransactionModel.gl_account
                )
            ),
            func.count(
                func.distinct(
                    TransactionModel.cost_center
                )
            )

        ).first()

        total_expense = float(summary[0] or 0)
        total_transaction = summary[1] or 0
        total_gl = summary[2] or 0
        total_cc = summary[3] or 0

        # Hitung average daily expense dari rentang actual data yang ter-query
        actual_range = summary_query.with_entities(
            func.min(TransactionModel.posting_date),
            func.max(TransactionModel.posting_date)
        ).first()

        actual_first = actual_range[0] if actual_range else None
        actual_last  = actual_range[1] if actual_range else None

        if actual_first and actual_last:
            total_days = count_workdays(actual_first, actual_last)
        else:
            total_days = 1

        average_daily_expense = (
            total_expense / total_days
            if total_days > 0
            else 0
        )

        # TOP GL ACCOUNT
        gl_query = (
            db.query(
                TransactionModel.gl_account,
                GlAccount.nama_gl_account,
                func.sum(TransactionModel.amount).label("total_amount")
            )
            .outerjoin(
                GlAccount,
                TransactionModel.gl_account == GlAccount.gl_account
            )
        )

        gl_query = apply_date_filter(
            gl_query,
            TransactionModel,
            start_date,
            end_date
        )

        gl = (
            gl_query
            .group_by(
                TransactionModel.gl_account,
                GlAccount.nama_gl_account
            )
            .order_by(
                func.sum(TransactionModel.amount).desc()
            )
            .limit(10)
            .all()
        )

        # TOP COST CENTER
        cc_query = db.query(
            TransactionModel.cost_center,

            func.sum(TransactionModel.amount).label("total_amount")
        )
        cc_query = apply_date_filter(
            cc_query,
            TransactionModel,
            start_date,
            end_date
        )

        cc = (
            cc_query
            .group_by(TransactionModel.cost_center)
            .order_by(
                func.sum(TransactionModel.amount).desc()
            )
            .limit(10)
            .all()
        )

        # TOP COST CENTER DETAIL
        top_cc = (
            db.query(
                TransactionModel.cost_center,
                func.sum(TransactionModel.amount).label("total_amount")
            )
        )

        top_cc = apply_date_filter(
            top_cc,
            TransactionModel,
            start_date,
            end_date
        )

        top_cc = (
            top_cc
            .group_by(TransactionModel.cost_center)
            .order_by(
                func.sum(TransactionModel.amount).desc()
            )
            .first()
        )

        detail = []

        if top_cc:
            detail_query = (
                db.query(
                    TransactionModel.gl_account,
                    GlAccount.nama_gl_account,
                    func.sum(TransactionModel.amount).label("total_amount")
                )
                .outerjoin(
                    GlAccount,
                    TransactionModel.gl_account == GlAccount.gl_account
                )
                .filter(
                    TransactionModel.cost_center == top_cc.cost_center
                )
            )

            detail_query = apply_date_filter(
                detail_query,
                TransactionModel,
                start_date,
                end_date
            )

            detail = (
                detail_query
                .group_by(
                    TransactionModel.gl_account,
                    GlAccount.nama_gl_account
                )
                .order_by(
                    func.sum(TransactionModel.amount).desc()
                )
                .all()
            )

        # TREND — ikuti trend_group dari parameter request
        # Validasi nilai trend_group
        if trend_group not in ("day", "month", "year"):
            trend_group = "month"

        if trend_group == "day":
            trend_query = db.query(
                func.date(TransactionModel.posting_date).label("period"),
                func.sum(TransactionModel.amount).label("total_amount")

            )

            trend_query = apply_date_filter(
                trend_query,
                TransactionModel,
                start_date,
                end_date
            )

            trend = (
                trend_query
                .group_by(
                    func.date(TransactionModel.posting_date)
                )
                .order_by(
                    func.date(TransactionModel.posting_date)
                )
                .all()
            )

        else:

            trend_query = db.query(
                TransactionModel.year,
                TransactionModel.month,
                func.sum(TransactionModel.amount).label("total_amount")
            )
            trend_query = apply_date_filter(
                trend_query,
                TransactionModel,
                start_date,
                end_date
            )
            trend = (
                trend_query
                .group_by(
                    TransactionModel.year,
                    TransactionModel.month
                )
                .order_by(
                    TransactionModel.year,
                    TransactionModel.month
                )
                .all()
            )
 
        # BUILD PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm
        )

        styles = getSampleStyleSheet()
        story = []

        # HEADER
        story.append(
            Paragraph(
                f"<b>Report Petty Cash {source.value.title()}</b>",
                styles["Title"]
            )
        )
        story.append(Spacer(1, 0.4 * cm))
        story.append(
            Paragraph(
                f"<b>Period :</b> {format_period(start_date, end_date)}",
                styles["Normal"]
            )
        )

        story.append(
            Paragraph(
                f"<b>Generated :</b> {datetime.now().strftime('%d-%m-%Y %H:%M')} WIB",
                styles["Normal"]
            )
        )

        story.append(Spacer(1, 0.6 * cm))

        # DASHBOARD SUMMARY
        story.append(
            Paragraph(
                "<b>Dashboard Summary</b>",
                styles["Heading2"]
            )
        )

        te_label, te_value = rupiah_parts(total_expense)
        ade_label, ade_value = rupiah_parts(average_daily_expense)
        summary_table = [
            ["Metric", "Value", ""],
            ["Total Expense", te_label, te_value],
            ["Total Transaction", "", f"{total_transaction:,}"],
            ["Total Cost Center", "", f"{total_cc:,}"],
            ["Average Daily Expense", ade_label, ade_value],
        ]
        story.append(
            create_table(
                summary_table,
                widths=[10 * cm, 1.0 * cm, 7 * cm],
                extra_style=[
                    ("SPAN", (1, 0), (2, 0)),               # header "Value" jadi satu
                    ("ALIGN", (1, 1), (1, -1), "LEFT"),     # "Rp." rata kiri
                    ("ALIGN", (2, 1), (2, -1), "RIGHT"),    # angka rata kanan
                    ("LINEAFTER", (1, 0), (1, -1), 0, colors.white),  # sembunyikan garis internal
                ]
            )
        )

        story.append(
            Spacer(1, 0.7 * cm)
        )

        # TOP GL ACCOUNT
        story.append(
            Paragraph(
                "<b>Top 10 GL Account</b>",
                styles["Heading2"]
            )
        )

        gl_table = [[
            "GL Account",
            "GL Name",
            "Amount",
            "",
            "%"
        ]]
        for row in gl:
            percentage = calculate_percentage(
                row.total_amount,
                total_expense
            )
            rp_label, rp_value = rupiah_parts(row.total_amount)
            gl_table.append([
                row.gl_account,
                row.nama_gl_account or "-",
                rp_label,
                rp_value,
                f"{percentage:.2f}%"
            ])
        story.append(
            create_table(
                gl_table,
                widths=[
                    3.0 * cm,   # GL Account
                    9.0 * cm,   # GL Name
                    0.8 * cm,   # "Rp."
                    3.2 * cm,   # Angka Amount
                    2.0 * cm,   # %
                ],
                extra_style=[
                    ("SPAN", (2, 0), (3, 0)),               # header "Amount" jadi satu
                    ("ALIGN", (2, 1), (2, -1), "LEFT"),     # "Rp." rata kiri
                    ("ALIGN", (3, 1), (3, -1), "RIGHT"),    # angka rata kanan
                    ("ALIGN", (4, 1), (4, -1), "CENTER"),   # % tidak rata kanan
                    ("LINEAFTER", (2, 0), (2, -1), 0, colors.white),
                ]
            )
        )

        story.append(
            Spacer(1,0.7*cm)
        )

        story.append(
            PageBreak()
        )
        # TOP COST CENTER
  
        story.append(
            Paragraph(
                "<b>Top 10 Cost Center</b>",
                styles["Heading2"]
            )
        )

        cc_table = [[
            "Cost Center",
            "Amount",
            "",
            "%"
        ]]
        for row in cc:
            percentage = calculate_percentage(
                row.total_amount,
                total_expense
            )
            rp_label, rp_value = rupiah_parts(row.total_amount)
            cc_table.append([
                row.cost_center,
                rp_label,
                rp_value,
                f"{percentage:.2f}%"
            ])
        story.append(
            create_table(
                cc_table,
                widths=[
                    10.0 * cm,  # Cost Center
                    0.8 * cm,   # "Rp."
                    4.2 * cm,   # Angka Amount
                    3.0 * cm,   # %
                ],
                extra_style=[
                    ("SPAN", (1, 0), (2, 0)),
                    ("ALIGN", (1, 1), (1, -1), "LEFT"),
                    ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                    ("ALIGN", (3, 1), (3, -1), "CENTER"),
                    ("LINEAFTER", (1, 0), (1, -1), 0, colors.white),
                ]
            )
        )

        # EXPENSE DETAIL BY TOP COST CENTER

        story.append(
            Paragraph(
                "<b>Expense Detail by Top Cost Center</b>",
                styles["Heading2"]
            )
        )

        if top_cc:

            story.append(
                Paragraph(
                    f"<b>Cost Center :</b> {top_cc.cost_center}",
                    styles["Normal"]
                )
            )

            story.append(
                Spacer(1,0.3*cm)
            )

            detail_table = [[
                "GL Account",
                "GL Name",
                "Amount",
                "",
                "%"
            ]]
            for row in detail:
                percentage = calculate_percentage(
                    row.total_amount,
                    top_cc.total_amount
                )
                rp_label, rp_value = rupiah_parts(row.total_amount)
                detail_table.append([
                    row.gl_account,
                    row.nama_gl_account or "-",
                    rp_label,
                    rp_value,
                    f"{percentage:.2f}%"
                ])
            story.append(
                create_table(
                    detail_table,
                    widths=[
                        3.0 * cm,   # GL Account
                        9.0 * cm,   # GL Name
                        0.8 * cm,   # "Rp."
                        3.2 * cm,   # Angka Amount
                        2.0 * cm,   # %
                    ],
                    extra_style=[
                        ("SPAN", (2, 0), (3, 0)),
                        ("ALIGN", (2, 1), (2, -1), "LEFT"),
                        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                        ("ALIGN", (4, 1), (4, -1), "CENTER"),
                        ("LINEAFTER", (2, 0), (2, -1), 0, colors.white),
                    ]
                )
            )

        else:

            story.append(
                Paragraph(
                    "No data available.",
                    styles["Normal"]
                )
            )

        story.append(
            Spacer(1,0.7*cm)
        )

        story.append(
            PageBreak()
        )
        
        # EXPENSE TREND
        story.append(
            Paragraph(
                "<b>Expense Trend</b>",
                styles["Heading2"]
            )
        )
        trend_table = [[
            "Period",
            "Amount",
            "",
            "Growth (%)"
        ]]
        previous_amount = None
        for row in trend:
            if trend_group == "day":
                period = str(row.period)
            else:
                period = f"{row.year}-{str(row.month).zfill(2)}"
            growth = calculate_growth(
                float(row.total_amount),
                previous_amount
            )
            rp_label, rp_value = rupiah_parts(row.total_amount)
            trend_table.append([
                period,
                rp_label,
                rp_value,
                "-" if growth is None else f"{growth:.2f}%"
            ])
            previous_amount = float(row.total_amount)
        story.append(
            create_table(
                trend_table,
                widths=[
                    5.0 * cm,   # Period
                    0.8 * cm,   # "Rp."
                    8.2 * cm,   # Angka Amount
                    4.0 * cm,   # Growth
                ],
                extra_style=[
                    ("SPAN", (1, 0), (2, 0)),
                    ("ALIGN", (1, 1), (1, -1), "LEFT"),
                    ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                    ("ALIGN", (3, 1), (3, -1), "CENTER"),
                    ("LINEAFTER", (1, 0), (1, -1), 0, colors.white),
                ]
            )
        )

        #  FOOTER
        story.append(
            Spacer(1,0.7*cm)
        )
        story.append(
            Paragraph(
                "<font size='8' color='grey'>Generated automatically by Navicash Dashboard.</font>",
                styles["Normal"]
            )
        )

        # BUILD PDF

        doc.build(story)
        buffer.seek(0)
        filename = (
            f"Report Petty Cash{source.value.title()}.pdf"
        )
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition":
                f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )
    
# Export Settlement Excel
@router.get("/settlement")
def export_settlement(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    employee_name: Optional[str] = Query(None),
    cost_center: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    try:
        query = db.query(Settlement).filter(
            Settlement.is_deleted == False
        )

        # Filter Tanggal
        if start_date:
            query = query.filter(
                Settlement.settlement_date >= start_date
            )

        if end_date:
            query = query.filter(
                Settlement.settlement_date <= end_date
            )

        # Filter Nama User
        if employee_name:
            query = query.filter(
                Settlement.employee_name.ilike(f"%{employee_name}%")
            )

        # Filter Cost Center
        if cost_center:
            query = query.filter(
                Settlement.cost_center.ilike(f"%{cost_center}%")
            )

        # Filter Source
        if source and source != "All Source":
            db_source = "ADVANCE" if source == "Settlement" else "REIMBURSEMENT"
            query = query.filter(
                Settlement.source == db_source
            )

        settlements = (
            query
            .order_by(
                Settlement.settlement_date.desc()
            )
            .all()
        )

        if not settlements:
            raise HTTPException(
                status_code=404,
                detail="Tidak ada data Settlement."
            )
        
        rows = []
        for item in settlements:
            rows.append({
                "Settlement Date": item.settlement_date,
                "PPC No": item.ppc_no,
                "Employee Name": item.employee_name,
                "Cost Center": item.cost_center,
                "Description": item.description or "",
                "Settlement Amount": item.settlement_amount,
                "Currency": "IDR",
                "Source": item.source.value if hasattr(item.source, "value") else str(item.source),
                "Checked": "Yes" if item.is_checked else "No",
            })

        df = pd.DataFrame(rows)
        output = BytesIO()

        with pd.ExcelWriter(
            output,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                index=False,
                sheet_name="Settlement"
            )

            # Auto fit column width
            worksheet = writer.sheets["Settlement"]
            format_excel_worksheet(
                worksheet,
                date_columns=[
                    "Settlement Date"
                ],
                currency_columns=[
                    "Settlement Amount"
                ]
            )

        output.seek(0)
        
        filename = f"settlement_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition":
                f'attachment; filename="{filename}"'
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )

# EXPORT ADVANCE
@router.get("/advance")
def export_ppc(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    employee_name: Optional[str] = Query(None),
    cost_center: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):

    query = db.query(AdvanceRequest)

    # Filter tanggal
    query = apply_advance_date_filter(
        query=query,
        start_date=start_date,
        end_date=end_date,
    )

    # Filter Nama User
    if employee_name:
        query = query.filter(
            AdvanceRequest.employee_name.ilike(f"%{employee_name}%")
        )

    # Filter Cost Center
    if cost_center:
        query = query.filter(
            AdvanceRequest.cost_center.ilike(f"%{cost_center}%")
        )

    # Filter Status
    if status and status != "All Status":
        status_map = {
            "Active": "ACTIVE",
            "Settled": "SETTLED",
            "Overdue": "OVERDUE",
            "Canceled": "CANCEL"
        }
        db_status = status_map.get(status)
        if db_status:
            query = query.filter(
                AdvanceRequest.status == db_status
            )

    ppc_list = (
        query
        .order_by(
            AdvanceRequest.request_date.desc()
        )
        .all()
    )

    # Tidak ada data
    if not ppc_list:
        raise HTTPException(
            status_code=404,
            detail="Tidak ada data PPC."
        )

    rows = []

    for ppc in ppc_list:

        # Update status terlebih dahulu
        update_ppc_status(ppc)

        status = ppc.status.value

        rows.append(
            {
                "Request Date": ppc.request_date,
                "PPC No": ppc.ppc_no,
                "Employee Name": ppc.employee_name,
                "Cost Center": ppc.cost_center,
                "Purpose": ppc.purpose,
                "Amount": ppc.amount,
                "Currency": "IDR",
                "Due Date": ppc.due_date,
                "Status": status,
            }
        )

    # Simpan perubahan status apabila ada yang berubah
    db.commit()

    df = pd.DataFrame(rows)

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name="Advance",
        )

        # Auto fit column width
        worksheet = writer.sheets["Advance"]

        format_excel_worksheet(
            worksheet,
            date_columns=[
                "Request Date",
                "Due Date"
            ],
            currency_columns=[
                "Amount"
            ]
        )

    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            "attachment; filename=advance_export.xlsx"
        },
    )

# EXPORT CHECK
@router.get("/check")
def export_check(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    transaction_type: Optional[str] = Query(None),
    bank_type: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):

    query = db.query(
        PrintedCheck
    )

    if start_date:
        query = query.filter(
            PrintedCheck.transaction_date >=
            start_date
        )

    if end_date:
        query = query.filter(
            PrintedCheck.transaction_date <=
            end_date
        )

    if transaction_type:
        query = query.filter(
            PrintedCheck.transaction_type ==
            transaction_type
        )

    if bank_type:
        clean_bank = bank_type.replace("Bank ", "").strip()
        query = query.filter(
            PrintedCheck.bank_type == clean_bank
        )

    checks = (
        query
        .order_by(
            PrintedCheck.updated_at.desc(),
            PrintedCheck.id.desc()
        )
        .all()
    )

    if not checks:
        raise HTTPException(
            status_code=404,
            detail=
            "Tidak ada data Check."
        )

    rows = []

    for item in checks:
        rows.append({
            "Tanggal": item.transaction_date,
            "Bank": f"Bank {item.bank_type.value}" if item.bank_type else "",
            "Nomor Cek": item.check_number,
            "Nominal": item.amount,
            "Currency": "IDR",
            "Vendor": item.vendor_name,
            "Nomor Rekening": item.vendor_account_number or "-",
            "Status": item.transaction_type.value if item.transaction_type else "",
        })

    df = pd.DataFrame(
        rows
    )
    output = BytesIO()
    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:
        df.to_excel(
            writer,
            sheet_name="History Cek",
            index=False
        )

        worksheet = writer.sheets["History Cek"]

        format_excel_worksheet(
            worksheet,
            date_columns=[
                "Tanggal"
            ],
            currency_columns=[
                "Nominal"
            ]
        )
        
    output.seek(0)
    filename = "export_historycek.xlsx"

    return StreamingResponse(
        output,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            f'attachment; filename="{filename}"'
        }
    )

# EXPORT VENDOR
@router.get("/vendor")
def export_vendor(
    db: Session = Depends(get_db)
):
    try:
        vendors = db.query(Vendor).order_by(Vendor.vendor_name.asc()).all()

        if not vendors:
            raise HTTPException(
                status_code=404,
                detail="Tidak ada data Vendor."
            )

        rows = []
        for item in vendors:
            rows.append({
                "Nama Vendor": item.vendor_name,
                "Nama Bank": item.bank_name,
                "Nama Akun Bank": item.bank_account_name,
                "No Rekening": item.bank_account_no,
            })

        df = pd.DataFrame(rows)
        output = BytesIO()

        with pd.ExcelWriter(
            output,
            engine="openpyxl"
        ) as writer:
            df.to_excel(
                writer,
                sheet_name="Vendor",
                index=False
            )

            worksheet = writer.sheets["Vendor"]
            format_excel_worksheet(
                worksheet
            )

        output.seek(0)
        filename = f"vendor {datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition":
                f'attachment; filename="{filename}"'
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )